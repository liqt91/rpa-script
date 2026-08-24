"""读取Excel内容 — excelRead (backend)"""
from pathlib import Path

from src.runtime.workflow.handlers.registry import register_handler, Param
from src.runtime.workflow.handlers.utils import clean_var_ref


@register_handler(
    cmd="excelRead", label="读取Excel内容", category="数据处理", runtime="backend",
    icon="fa-file-excel", icon_color="text-green-500", bg_color="bg-green-50",
    category_order=45, command_order=10,
    description="读取Excel(.xlsx)单元格或区域的内容，保存到变量。单个单元格返回标量，区域或整表返回二维列表。",
    summary_tpl="{filePath}")
class ExcelReadHandler:
    params = [
        Param("filePath", "Excel文件路径", "string", required=True,
              placeholder="如 C:\\data\\report.xlsx"),
        Param("sheet", "工作表名", "string", placeholder="留空用第一个工作表"),
        Param("range", "单元格/区域", "string",
              placeholder="如 A1 单个单元格，或 A1:C10 区域；留空读取整张表"),
        Param("resultVar", "保存到变量", "str-var", default="excelData", group="output"),
    ]

    @staticmethod
    async def execute(runner, cmd_type, step_id, instr):
        extra = instr.get("extra", {})
        file_path = (extra.get("filePath") or "").strip()
        sheet = (extra.get("sheet") or "").strip()
        range_ref = (extra.get("range") or "").strip()
        result_var = clean_var_ref(extra.get("resultVar", "excelData"))

        if not file_path:
            raise ValueError("Excel文件路径不能为空")
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel文件不存在: {file_path}")

        try:
            import openpyxl
            from openpyxl.utils.cell import range_boundaries
        except ImportError as e:
            raise RuntimeError(
                "读取Excel需要 openpyxl，请先安装：pip install openpyxl") from e

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            raise RuntimeError(f"无法打开Excel文件 {file_path}: {e}") from e

        try:
            if sheet:
                if sheet not in wb.sheetnames:
                    raise ValueError(f"找不到工作表: {sheet}（可用: {wb.sheetnames}）")
                ws = wb[sheet]
            else:
                ws = wb.worksheets[0]

            target = range_ref or ws.calculate_dimension()
            data = ExcelReadHandler._read_region(ws, target, range_boundaries)
        finally:
            wb.close()

        result = {
            "excelRead": file_path,
            "sheet": ws.title,
            "range": target,
            "value": data,
        }
        runner.completed += 1
        runner.vars[result_var] = data
        runner.results.append({
            "stepId": step_id,
            "nodeId": instr.get("nodeId"),
            "status": "success",
            "result": result,
        })
        await runner._emit({
            "type": "stepComplete",
            "stepId": step_id,
            "nodeId": instr.get("nodeId"),
            "result": result,
        })
        return True

    @staticmethod
    def _read_region(ws, target, range_boundaries):
        """读取单元格(单格→标量)或区域(→二维列表)。"""
        min_col, min_row, max_col, max_row = range_boundaries(target)
        if min_col == max_col and min_row == max_row:
            return ExcelReadHandler._jsonable(ws.cell(row=min_row, column=min_col).value)
        rows = []
        for row in ws.iter_rows(
                min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
            rows.append([ExcelReadHandler._jsonable(c.value) for c in row])
        return rows

    @staticmethod
    def _jsonable(v):
        """把日期/时间等无法直接JSON序列化的值转成可序列化形式。"""
        import datetime as _dt
        if isinstance(v, (_dt.datetime, _dt.date, _dt.time)):
            return v.isoformat()
        return v
